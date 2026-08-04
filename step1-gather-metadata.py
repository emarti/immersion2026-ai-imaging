#!/usr/bin/env python3
"""Step 1: gather per-subject metadata from the OASIS-1 dataset.

Walks every ``disc*/OAS1_XXXX_MRy/OAS1_XXXX_MRy.txt`` under the configured data
root, parses age / sex / CDR (and MMSE for reference), locates the matching
brain-masked atlas volume, derives the binary dementia label, and writes a
tidy ``metadata.csv``. It also saves ``outputs/step1-dataset_age_histograms.png`` --
CDR-negative vs CDR-positive counts by age over the *entire* dataset (step2 makes the same
plot restricted to the cohort's age range).

Label rule (from the OASIS fact sheet):
    CDR == 0            -> 0  (CDR-negative)
    CDR in {0.5, 1, 2}  -> 1  (CDR-positive; any impairment, clinically probable AD)
    CDR blank           -> left empty (young subjects were not assessed;
                                       step2 excludes these rows)

Usage:
    python step1-gather-metadata.py [--config config.yaml]
"""
from __future__ import annotations

import argparse
import csv
import glob
import os

from common import load_config, metadata_csv, reference_xlsx


def parse_txt(txt_path: str) -> dict:
    """Extract the fields we care about from a session .txt file.

    The file is a simple ``KEY:   value`` listing. We only read the header
    block (age / sex / CDR / MMSE); later scan-parameter blocks reuse some key
    names, so we stop at the first blank line.
    """
    fields: dict = {}
    with open(txt_path, "r", errors="ignore") as f:
        for line in f:
            if line.strip() == "":
                break  # header block ends at the first blank line
            if ":" not in line:
                continue
            key, _, value = line.partition(":")
            fields[key.strip().upper()] = value.strip()
    return fields


def to_float(value: str):
    """Parse a numeric string, returning None for blanks / 'N/A'."""
    if value is None:
        return None
    value = value.strip()
    if value == "" or value.upper() == "N/A":
        return None
    try:
        return float(value)
    except ValueError:
        return None


def normalize_sex(value: str):
    if not value:
        return None
    v = value.strip().upper()
    if v.startswith("M"):
        return "M"
    if v.startswith("F"):
        return "F"
    return None


def find_image(session_dir: str, image_glob: str):
    matches = sorted(glob.glob(os.path.join(session_dir, image_glob)))
    return matches[0] if matches else None


def gather(config: dict) -> list[dict]:
    data_root = config["data_raw_path"]
    image_glob = config["image_glob"]
    rows: list[dict] = []

    for disc in config["discs"]:
        disc_dir = os.path.join(data_root, disc)
        if not os.path.isdir(disc_dir):
            print(f"  [skip] disc not found on disk: {disc_dir}")
            continue

        for session in sorted(os.listdir(disc_dir)):
            session_dir = os.path.join(disc_dir, session)
            txt_path = os.path.join(session_dir, f"{session}.txt")
            if not os.path.isfile(txt_path):
                continue

            fields = parse_txt(txt_path)
            cdr = to_float(fields.get("CDR"))
            age = to_float(fields.get("AGE"))
            sex = normalize_sex(fields.get("M/F"))
            mmse = to_float(fields.get("MMSE"))

            img_path = find_image(session_dir, image_glob)
            label = "" if cdr is None else (1 if cdr > 0 else 0)

            rows.append(
                {
                    "subject": session,
                    "disc": disc,
                    "age": "" if age is None else int(age),
                    "sex": sex or "",
                    "cdr": "" if cdr is None else cdr,
                    "mmse": "" if mmse is None else mmse,
                    "label": label,
                    "img_path": img_path or "",
                    "img_exists": bool(img_path),
                }
            )
    return rows


def write_csv(rows: list[dict], out_path: str) -> None:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    columns = [
        "subject", "disc", "age", "sex", "cdr", "mmse",
        "label", "img_path", "img_exists",
    ]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def summarize(rows: list[dict]) -> None:
    total = len(rows)
    with_cdr = sum(1 for r in rows if r["label"] != "")
    cdr_positive = sum(1 for r in rows if r["label"] == 1)
    cdr_negative = sum(1 for r in rows if r["label"] == 0)
    with_img = sum(1 for r in rows if r["img_exists"])
    in70s = sum(1 for r in rows if r["age"] != "" and 70 <= int(r["age"]) <= 79
                and r["label"] != "")
    print("\nSummary")
    print(f"  sessions found ............. {total}")
    print(f"  with CDR (labelled) ........ {with_cdr}  (CDR+={cdr_positive}, CDR-={cdr_negative})")
    print(f"  without CDR (excluded) ..... {total - with_cdr}")
    print(f"  with masked_gfc image ...... {with_img}")
    print(f"  labelled subjects in 70s ... {in70s}")


def _num(value):
    """Coerce a spreadsheet or metadata value to float, or None for blank/'N/A'.

    openpyxl yields numbers as int/float and blanks as None; the metadata rows
    hold ints/floats or ''. ``to_float`` covers the string path (including 'N/A').
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return to_float(value)


# metadata.csv field -> (reference column, comparison normalizer)
COMPARE_FIELDS = {
    "sex": ("M/F", normalize_sex),
    "age": ("Age", _num),
    "cdr": ("CDR", _num),
    "mmse": ("MMSE", _num),
}


def load_reference(xlsx_path: str) -> dict:
    """Read the OASIS cross-sectional spreadsheet -> {subject_id: row dict}."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    out: dict = {}
    for r in it:
        rec = dict(zip(header, r))
        sid = rec.get("ID")
        if sid:
            out[str(sid).strip()] = rec
    wb.close()
    return out


def validate_against_reference(rows: list[dict], xlsx_path: str) -> None:
    """Cross-check metadata.csv rows against the reference spreadsheet.

    Compares only the overlapping fields (sex, age, CDR, MMSE); the spreadsheet's
    Educ/SES/eTIV/nWBV/ASF/Delay have no metadata.csv counterpart. Reports every
    discrepancy but never raises -- validation is advisory, not a gate.
    """
    print("\nValidation vs reference spreadsheet")
    if not os.path.isfile(xlsx_path):
        print(f"  [skip] reference not found: {xlsx_path}")
        return
    try:
        ref = load_reference(xlsx_path)
    except ImportError:
        print("  [skip] openpyxl not installed (add it: mamba install -c conda-forge openpyxl)")
        return
    except Exception as e:  # noqa: BLE001
        print(f"  [skip] could not read {xlsx_path}: {e}")
        return

    meta = {r["subject"]: r for r in rows}
    only_meta = sorted(set(meta) - set(ref))
    only_ref = sorted(set(ref) - set(meta))

    diffs = {field: [] for field in COMPARE_FIELDS}
    for sid in sorted(set(meta) & set(ref)):
        m, x = meta[sid], ref[sid]
        for field, (col, norm) in COMPARE_FIELDS.items():
            mv, xv = norm(m.get(field, "")), norm(x.get(col))
            if mv != xv:
                diffs[field].append((sid, m.get(field, ""), x.get(col)))

    print(f"  reference subjects .. {len(ref)}")
    print(f"  matched subjects .... {len(set(meta) & set(ref))}")
    if only_meta:
        print(f"  only in metadata ({len(only_meta)}): {', '.join(only_meta)}")
    if only_ref:
        print(f"  only in reference ({len(only_ref)}): {', '.join(only_ref)}")

    total_diffs = sum(len(v) for v in diffs.values())
    for field, entries in diffs.items():
        if entries:
            print(f"  {field}: {len(entries)} mismatch(es) [metadata != reference]")
            for sid, mv, xv in entries:
                print(f"    {sid}: {mv!r} != {xv!r}")

    if not only_meta and not only_ref and total_diffs == 0:
        print(f"  OK -- matches reference ({len(ref)} subjects)")
    else:
        print(f"  {len(only_meta) + len(only_ref) + total_diffs} discrepancy(ies) found (see above)")


def plot_age_histograms(rows: list[dict], out_path: str, labels_cfg: dict) -> None:
    """Save a 3-panel age histogram of CDR-negative vs CDR-positive over the ENTIRE dataset.

    (Deliberately duplicated from step2's near-identical plot -- this teaching project keeps the
    two steps self-contained. step1 shows every labelled subject at any age; step2 restricts the
    same plot to the cohort's age_min..age_max range.) Panels: (1) both sexes pooled; (2) sexes
    separated (line style); (3) raw CDR grade (0 / 0.5 / 1 / 2) separated.

    Note: step1's rows hold native types (int age, int label, bool img_exists, float cdr),
    unlike step2 which reads strings back from metadata.csv.
    """
    import matplotlib
    matplotlib.use("Agg")                 # write a file, no interactive window
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator
    import numpy as np

    plt.style.use("seaborn-v0_8-darkgrid")

    recs = []
    for r in rows:
        if r["label"] not in (0, 1):                 # native ints ("" excluded)
            continue
        if not r["img_exists"] or r["age"] == "" or r["sex"] not in ("M", "F"):
            continue
        try:
            age, cdr = int(r["age"]), float(r["cdr"])
        except (ValueError, TypeError):
            continue
        recs.append((age, r["sex"], int(r["label"]), cdr))
    if not recs:
        print("  [skip] no labelled subjects with images to plot age histograms")
        return

    ages = np.array([a for a, _, _, _ in recs])
    sexes = np.array([s for _, s, _, _ in recs])
    labels = np.array([l for _, _, l, _ in recs])
    cdrs = np.array([c for _, _, _, c in recs])

    edges = np.arange((ages.min() // 5) * 5, ((ages.max() // 5) + 1) * 5 + 1, 5)  # 5-yr bins
    neg_name = labels_cfg.get("cdr_negative", "CDR Negative")
    pos_name = labels_cfg.get("cdr_positive", "CDR Positive")
    C_NEG, C_POS = "steelblue", "crimson"          # CDR- = blue / CDR+ = red, as elsewhere

    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(16, 4.6), sharex=True)

    # (1) both sexes pooled
    ax1.hist(ages[labels == 0], bins=edges, histtype="stepfilled", alpha=0.55,
             color=C_NEG, label=neg_name)
    ax1.hist(ages[labels == 1], bins=edges, histtype="stepfilled", alpha=0.55,
             color=C_POS, label=pos_name)
    ax1.set(title="CDR status vs age (both sexes)", xlabel="age (years)", ylabel="subjects")
    ax1.legend(fontsize=8)

    # (2) sexes separated (colour = CDR, solid = Male / dashed = Female)
    for lbl, colour, name in ((0, C_NEG, neg_name), (1, C_POS, pos_name)):
        for sex, ls, sname in (("M", "-", "Male"), ("F", "--", "Female")):
            m = (labels == lbl) & (sexes == sex)
            if m.any():
                ax2.hist(ages[m], bins=edges, histtype="step", linewidth=1.6,
                         color=colour, linestyle=ls, label=f"{name} · {sname}")
    ax2.set(title="CDR status vs age (sexes separated)", xlabel="age (years)")
    ax2.legend(fontsize=7)

    # (3) raw CDR grade separated (both sexes pooled)
    grade_colours = {0.0: "steelblue", 0.5: "gold", 1.0: "darkorange", 2.0: "crimson"}
    for g in (0.0, 0.5, 1.0, 2.0):
        m = cdrs == g
        if m.any():
            ax3.hist(ages[m], bins=edges, histtype="step", linewidth=1.8,
                     color=grade_colours[g], label=f"CDR {g:g}")
    ax3.set(title="CDR grade vs age (both sexes)", xlabel="age (years)")
    ax3.legend(fontsize=8)

    for ax in (ax1, ax2, ax3):                      # integer counts -> integer y ticks
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))

    fig.suptitle("OASIS-1 age distribution by CDR status -- entire dataset "
                 "(all labelled subjects with an image)", fontsize=11)
    fig.tight_layout(rect=(0, 0, 1, 0.96))
    fig.savefig(out_path, dpi=120)
    plt.close(fig)
    print(f"Saved age histograms -> {out_path}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--config", default="config.yaml")
    ap.add_argument("--skip-validate", action="store_true",
                    help="skip the metadata.csv vs reference-spreadsheet cross-check")
    args = ap.parse_args()

    config = load_config(args.config)
    print(f"Scanning {config['data_raw_path']} across discs: {', '.join(config['discs'])}")
    rows = gather(config)

    out_path = metadata_csv(config)
    write_csv(rows, out_path)
    summarize(rows)
    print(f"\nWrote {len(rows)} rows -> {out_path}")

    outputs_path = config["outputs_path"]
    os.makedirs(outputs_path, exist_ok=True)
    plot_age_histograms(rows, os.path.join(outputs_path, "step1-dataset_age_histograms.png"),
                        config.get("labels") or {})

    if not args.skip_validate:
        validate_against_reference(rows, reference_xlsx(config))


if __name__ == "__main__":
    main()
